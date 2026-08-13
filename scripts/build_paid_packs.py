#!/usr/bin/env python3
"""Build the paid packs into content/packs/:

  starter-pack.zip   Diagnosis Prep Starter Pack (PDF, 7 sections)
  full-pack.zip      Full Diagnosis Prep & Survival Pack (PDF, 15 sections)
  admin-system.zip   ADHD Admin System (browser app + Excel + Notion guide)
  bundle.zip         Full Pack PDF + Admin System files

Usage:  python3 scripts/build_paid_packs.py
"""

import os
import shutil
import zipfile

from reportlab.platypus import HRFlowable, PageBreak, Paragraph, Spacer

from packlib import (
    CONTENT_W, CORAL, OUT_DIR, S, bullets, build_pdf, checks, cover,
    disclaimer_flow, grid_table, note_box, script_box, section_header,
    write_lines,
)

HERE = os.path.dirname(os.path.abspath(__file__))
SITE = "adhd-assist.vercel.app"


# ===========================================================================
# Shared section builders (used by Starter and Full — Full passes its own
# section numbers/totals)
# ===========================================================================

def sec_gp_prep(num, total):
    f = section_header(num, total, "Full GP Appointment Prep Template",
        "Work through Parts A–H before your appointment, then bring the whole thing. "
        "It doubles as your script if your mind goes blank.")
    f += [
        Paragraph("Part A — Why now?", S["h2"]),
        Paragraph("One or two sentences on what made you finally book this. "
                  "(“I read something and it all clicked” is a perfectly good answer.)", S["small"]),
        *write_lines(2),
        Paragraph("Part B — The history", S["h2"]),
        Paragraph("How far back does this go? School reports, teachers’ comments, family stories. "
                  "Assessors specifically look for evidence before age 12.", S["small"]),
        *write_lines(4),
        Paragraph("Part C — Current struggles (top 3, with recent examples)", S["h2"]),
        *write_lines(3, "1."),
        *write_lines(3, "2."),
        *write_lines(3, "3."),
        PageBreak(),
        Paragraph("Part D — More things that are true (tick what applies)", S["h2"]),
        *checks([
            "I lose focus even during things I care about",
            "I put off tasks until they become emergencies",
            "My systems (apps, planners, folders) collapse within weeks",
            "I interrupt or talk over people without meaning to",
            "I lose everyday items constantly",
            "Deadlines feel unreal until they are terrifying",
            "I have periods of intense focus where hours disappear",
            "Small criticisms hit me much harder than they should",
        ]),
        Paragraph("Part E — The impact", S["h2"]),
        Paragraph("What has this cost you? Work, study, money, relationships, driving, health, "
                  "self-esteem. Impact is what turns “quirks” into a referral.", S["small"]),
        *write_lines(4),
        Paragraph("Part F — My opening statement", S["h2"]),
        Paragraph("Write it, practise it once out loud, say it in the first 30 seconds.", S["small"]),
        *script_box("A shape that works",
            "I’ve struggled with focus, organisation and memory my whole life. It’s causing real "
            "problems with [work / money / relationships], and I’d like to be assessed for ADHD."),
        *write_lines(2),
        PageBreak(),
        Paragraph("Part G — Questions to ask before you leave", S["h2"]),
        *checks([
            "Where are you referring me, and roughly how long is the wait?",
            "Am I eligible for Right to Choose? (England only)",
            "How will I hear about the referral — and who do I chase if I hear nothing?",
            "Is there anything I should do or collect while I wait?",
        ]),
        Paragraph("Part H — Straight after the appointment", S["h2"]),
        *checks([
            "Write down what was agreed while it’s fresh (same hour if possible)",
            "Note the referral destination, date, and any reference given",
            "Diary a chase-up for 6–8 weeks if you’ve heard nothing",
        ]),
        *write_lines(3, "Notes from the appointment:"),
        PageBreak(),
    ]
    return f


EVIDENCE_CATS = [
    ("Focus & attention", "Drifting mid-task or mid-conversation, rereading the same page, careless mistakes in things you understand."),
    ("Hyperfocus", "Hours vanishing into one (often unhelpful) task; forgetting to eat, drink or stop."),
    ("Time management", "Chronic lateness, “time blindness”, the loads-of-time → suddenly-late flip."),
    ("Organisation", "Unopened post, chaotic inbox, systems that collapse, abandoned projects."),
    ("Memory", "Forgetting commitments minutes after making them; lost keys, phone, wallet; blank moments."),
    ("Emotional regulation", "Quick frustration, criticism that lands far too hard, big feelings that take hours to settle."),
    ("Impulsivity", "Spending, blurting, snap decisions — regretted almost immediately."),
    ("Restlessness", "Feeling driven by a motor; fidgeting; a mind that never idles, even when exhausted."),
]


def sec_evidence(num, total):
    f = section_header(num, total, "Full Symptom Evidence Builder",
        "Eight categories. For each: one or two recent examples, one older/childhood example, "
        "and what it costs you. Specific and dated beats vague and vivid.")
    for i, (name, hint) in enumerate(EVIDENCE_CATS):
        f += [
            Paragraph(f"{i + 1}. {name}", S["h2"]),
            Paragraph(hint, S["small"]),
            *write_lines(3),
        ]
        if i == 3:
            f.append(PageBreak())
    f += [
        Spacer(1, 4),
        Paragraph("Pattern summary", S["h2"]),
        Paragraph("Read back what you wrote. In one or two sentences: what’s the pattern, "
                  "and how long has it been there?", S["small"]),
        *write_lines(3),
        PageBreak(),
    ]
    return f


def sec_childhood(num, total):
    f = section_header(num, total, "Childhood & Adult Examples Worksheet",
        "Assessment criteria ask for evidence that traits were present in childhood. "
        "You don’t need perfect memories — patterns and stories count.")
    f += [
        Paragraph("Primary school (roughly ages 5–11)", S["h2"]),
        Paragraph("Report comments, lost belongings, daydreaming, trouble sitting still, "
                  "unfinished work, “bright but…”.", S["small"]),
        *write_lines(3),
        Paragraph("Secondary school (roughly ages 11–16)", S["h2"]),
        Paragraph("Homework chaos, last-minute cramming, detentions, forgotten kit, "
                  "grades that swung with interest.", S["small"]),
        *write_lines(3),
        Paragraph("16–18 and further study", S["h2"]),
        *write_lines(3),
        PageBreak(),
        Paragraph("Early adulthood", S["h2"]),
        Paragraph("First jobs, house shares, bills, driving, relationships — where did the "
                  "pattern follow you?", S["small"]),
        *write_lines(3),
        Paragraph("The “always” list", S["h2"]),
        Paragraph("Things people have said about you your whole life: “away with the fairies”, "
                  "“so much potential”, “last-minute merchant”…", S["small"]),
        *write_lines(3),
        Paragraph("Who could confirm the childhood picture?", S["h2"]),
        Paragraph("A parent, sibling, or old friend. Some providers ask for an informant "
                  "questionnaire — warn this person in advance.", S["small"]),
        *write_lines(2),
        *note_box("If you have old school reports, dig them out and bring copies. They are the "
                  "single most persuasive piece of childhood evidence."),
        PageBreak(),
    ]
    return f


def sec_scripts(num, total, include_rtc):
    n_scripts = 10 if include_rtc else 9
    f = section_header(num, total, f"Phone Scripts: Every Scenario ({n_scripts} scripts)",
        "Read them out word-for-word — nobody can tell, and nobody cares. "
        "Adapt the bracketed bits.")
    f += [
        *script_box("1 · Booking the appointment",
            "Hi, I’d like to book a GP appointment to discuss being assessed for ADHD. "
            "Do you have anything in the next couple of weeks?"),
        *script_box("2 · Asking for a double appointment",
            "It’s a conversation that won’t fit in ten minutes — could I book a double "
            "appointment, or with whichever GP has the longest slot?"),
        *script_box("3 · If the receptionist asks why",
            "It’s about a long-term difficulty with focus and organisation that’s affecting "
            "my daily life. I’d like to discuss an ADHD assessment with the GP."),
        *script_box("4 · If your mind goes blank mid-call",
            "Sorry — I’ve lost my thread, which is honestly part of the problem. Give me two "
            "seconds… I’m calling about booking an appointment for an ADHD assessment."),
        *script_box("5 · Online form / e-consult wording",
            "Lifelong difficulties with focus, organisation and memory, now causing significant "
            "problems at [work/home]. Requesting an appointment to discuss ADHD assessment "
            "and referral. Have prepared notes and examples."),
        PageBreak(),
        *script_box("6 · In the appointment, if you freeze",
            "Sorry — my mind’s gone blank, which is actually part of why I’m here. I’ve written "
            "everything down — can I read from my notes?"),
        *script_box("7 · If the GP seems dismissive",
            "I understand, but this has been a lifelong pattern and it’s seriously affecting my "
            "work and my health. I’d still like a referral for a proper assessment. Could we "
            "record that I’ve asked?"),
        *script_box("8 · Chasing a referral",
            "Hi, I’m chasing an ADHD assessment referral for [name, date of birth], referred by "
            "[surgery] on [date], reference [number if you have one]. Could you confirm it’s in "
            "the queue and give me a current estimated wait?"),
        *script_box("9 · Calling a private provider",
            "Hi, I’m considering an ADHD assessment with you. Could you tell me your current "
            "wait, the total cost including titration, and whether GP surgeries usually accept "
            "your shared-care agreements?"),
    ]
    if include_rtc:
        f += script_box("10 · Asking for a Right to Choose referral (England)",
            "I’d like to use my legal Right to Choose and be referred to [provider name] for my "
            "ADHD assessment. They hold an NHS contract, so the referral goes through the "
            "e-Referral system — I can leave their details with you.")
    f += [
        *note_box("You are allowed to read from notes. You are allowed to ask again. "
                  "None of this counts against you."),
        PageBreak(),
    ]
    return f


def sec_followup(num, total):
    f = section_header(num, total, "Appointment Follow-Up Checklist",
        "The hour after an appointment is where progress either gets banked or evaporates. "
        "Run this after every appointment — GP, provider, or review.")
    f += [
        Paragraph("Same hour", S["h2"]),
        *checks([
            "Write down what was said and agreed — three bullet points is enough",
            "Note any names, reference numbers, or timescales mentioned",
            "If anything was promised (“we’ll send…”, “you’ll hear within…”), note the date it’s due",
        ]),
        Paragraph("Within 48 hours", S["h2"]),
        *checks([
            "Diary the chase-up for anything promised, at its due date",
            "File any paperwork in one place (one folder, physical or digital)",
            "Tell whoever supports you what happened — saying it out loud helps it stick",
        ]),
        Paragraph("Within 2 weeks", S["h2"]),
        *checks([
            "Confirm the referral/letter actually went where it was meant to go",
            "Add any new examples that surfaced since the appointment to your evidence sheet",
        ]),
        Paragraph("If it didn’t go the way you hoped", S["h2"]),
        *bullets([
            "You can book another appointment with a different GP at the same surgery.",
            "You can put the request in writing (letter or e-consult) and ask for it to be recorded.",
            "A “no” from one appointment is not a diagnosis — it’s one conversation.",
        ]),
        *write_lines(3, "Notes:"),
        PageBreak(),
    ]
    return f


AI_PROMPTS_STARTER = [
    ("Overwhelm triage", "I have ADHD and I’m overwhelmed. Here’s everything on my plate: [brain-dump]. Ask me questions one at a time to find the single most urgent item, then break it into steps of five minutes or less, and tell me only the first step."),
    ("Turning chaos into GP notes", "Here’s a messy description of my struggles: [paste]. Rewrite it as three clear bullet points a GP can absorb in one minute, each with the pattern, one example, and the impact."),
    ("Finding childhood examples", "Interview me, one question at a time, about my school years — focus, homework, losing things, sitting still, teachers’ comments. Then summarise the ADHD-relevant patterns you noticed."),
    ("The opening statement", "Using these notes [paste], draft a two-sentence opening statement for my GP appointment asking for an ADHD assessment. Plain, direct, no drama."),
    ("Rehearsing the appointment", "Role-play a UK GP appointment with me. You’re a kind but time-pressed GP. Let me practise asking for an ADHD assessment, then give me feedback on what was clear and what was vague."),
    ("Decoding a letter", "Explain this NHS/provider letter in plain English: [paste]. What is it saying, what do they want from me, and what should I do next? Note anything time-sensitive."),
    ("Writing the e-consult", "Turn these notes [paste] into a short online-consultation message requesting an ADHD assessment referral: factual, specific, under 150 words."),
    ("Chasing without spiralling", "Draft a short, polite email chasing my ADHD referral (referred [date] by [surgery] to [provider]). Firm, friendly, asks for confirmation and a current wait estimate."),
    ("Post-appointment debrief", "Here’s what happened at my appointment: [paste]. Summarise what was agreed, list my action items with suggested dates, and flag anything I should chase if I don’t hear back."),
]

AI_PROMPTS_FULL_EXTRA = [
    ("Choosing a route", "Ask me one question at a time about my budget, how long I can wait, where I live in the UK, and how I feel about phone/video appointments. Then lay out NHS, Right to Choose and private routes for my situation, with the trade-offs."),
    ("Comparing providers", "Here are my notes on three ADHD providers: [paste]. Build a comparison table across wait, total cost, titration, shared care record, and reviews, then tell me what question I still need answered before choosing."),
    ("Right to Choose request", "Draft a short message to my GP requesting a Right to Choose referral to [provider] for ADHD assessment, mentioning that they hold an NHS contract and the referral goes via the e-Referral system."),
    ("Weekly survival plan", "I’m waiting for an ADHD assessment and struggling. Help me build a minimum viable week: sleep anchor, one priority per day, and one thing I’m allowed to drop. Keep it small enough to actually happen."),
    ("Referral limbo check-in", "It’s been [X weeks] since my referral to [provider]. Based on UK norms, is it time to chase? Draft the chase message, and tell me what to ask so the answer is actually useful."),
    ("Assessment day prep", "My ADHD assessment is on [date] with [provider]. Build me a prep sheet: what to bring, the five points I must not forget to say [paste notes], and questions to ask at the end."),
    ("Decoding the report", "Here’s the summary of my assessment report: [paste]. Explain it in plain English, list the recommendations, and tell me what to ask my GP and the provider next."),
    ("Not diagnosed — now what", "I wasn’t diagnosed with ADHD. Here’s what the report said: [paste]. Help me understand the reasoning, what alternative explanations were suggested, and what sensible next steps look like."),
    ("Shared care request", "Draft a polite letter to my GP surgery asking whether they will accept a shared-care agreement with [provider] for ADHD medication after titration — and, if not, asking for the reason in writing."),
    ("Titration diary summary", "Here are my daily titration notes: [paste]. Summarise into what my prescriber needs: overall trend, side effects and their timing, wear-off pattern, and three questions worth asking at review."),
    ("Medication review prep", "My medication review is on [date]. From these notes [paste], produce a one-page summary: what’s working, what isn’t, changes since last review, and what I want from this appointment."),
]


def sec_ai_prompts(num, total, full):
    if full:
        title = "Full AI Prompt Library (20 prompts)"
        sub = ("Copy-paste into ChatGPT, Claude, or Gemini. Organised by stage. "
               "Replace the [bracketed] parts with your own material.")
    else:
        title = "AI Prompt Pack — Appointment & Admin Edition (9 prompts)"
        sub = ("Copy-paste into ChatGPT, Claude, or Gemini. "
               "Replace the [bracketed] parts with your own material.")
    f = section_header(num, total, title, sub)
    prompts = AI_PROMPTS_STARTER + (AI_PROMPTS_FULL_EXTRA if full else [])
    for i, (label, text) in enumerate(prompts):
        f += script_box(f"{i + 1} · {label}", text)
        if i % 3 == 2:
            f.append(Spacer(1, 2))
    f += [
        *note_box("AI tools help you organise and rehearse — they can’t diagnose you or give "
                  "medical advice. Keep health decisions with qualified clinicians."),
        PageBreak(),
    ]
    return f


def sec_referral_docs(num, total):
    f = section_header(num, total, "Referral Document Checklist",
        "Everything worth gathering before and during the referral stage, "
        "so nothing has to be found twice.")
    f += [
        Paragraph("Gather (tick when it’s in your folder)", S["h2"]),
        *checks([
            "Your symptom evidence sheets (this pack, sections above)",
            "Old school reports or any childhood records you can find",
            "A list of current medications and doses",
            "Relevant medical history notes (other diagnoses, past mental-health support)",
            "Photo ID and NHS number (on prescriptions or the NHS app)",
        ]),
        Paragraph("Ask your GP for", S["h2"]),
        *checks([
            "Confirmation the referral has been sent (date + destination)",
            "A copy of the referral letter for your records",
            "Your NHS number if you don’t have it",
        ]),
        Paragraph("Contact tracker", S["h2"]),
        *grid_table(
            ["Date", "Who / organisation", "What was said", "Next step + who owes it"],
            [["", "", "", ""] for _ in range(8)],
            [CONTENT_W * 0.13, CONTENT_W * 0.24, CONTENT_W * 0.33, CONTENT_W * 0.30],
            row_h=24),
        Paragraph("Provider information record", S["h2"]),
        *grid_table(
            ["Field", "Details"],
            [["Provider name", ""], ["Contact phone / email", ""],
             ["Referral date & reference", ""], ["Estimated wait", ""],
             ["Next chase-up due", ""]],
            [CONTENT_W * 0.35, CONTENT_W * 0.65], row_h=22),
        PageBreak(),
    ]
    return f


# ===========================================================================
# Full-pack-only sections
# ===========================================================================

def sec_routes_full(num, total):
    f = section_header(num, total, "Understanding Your Routes",
        "The complete guide to the three UK routes — and how to pick.")
    f += [
        Paragraph("Route 1 — Standard NHS referral", S["h2"]),
        *bullets([
            "Free. GP refers you to the local NHS ADHD service.",
            "Waits in many areas run to years. Ask for the current local estimate.",
            "You can switch to Right to Choose or private later without starting over.",
        ]),
        Paragraph("Route 2 — Right to Choose (England only)", S["h2"]),
        *bullets([
            "NHS-funded and free — you choose any provider holding an NHS contract.",
            "Waits are often weeks-to-months. Your GP refers via the NHS e-Referral system.",
            "You must be registered with an English GP practice.",
            "Availability has been changing in some areas — check the current position "
            "(ADHD UK’s guide at adhduk.co.uk is kept up to date).",
        ]),
        Paragraph("Route 3 — Private", S["h2"]),
        *bullets([
            "Fastest. Typical assessment cost: several hundred pounds to £1,200+.",
            "Budget the whole journey: assessment + titration + prescriptions + annual reviews.",
            "Shared care is the pivot: if your GP won’t take it on, private prices continue indefinitely.",
        ]),
        Paragraph("Side by side", S["h2"]),
        *grid_table(
            ["", "NHS standard", "Right to Choose", "Private"],
            [
                ["Cost to you", "Free", "Free", "£££ (assessment + ongoing)"],
                ["Typical wait", "Often years", "Weeks–months", "Days–weeks"],
                ["Who chooses provider", "The NHS", "You", "You"],
                ["Where available", "UK-wide", "England only", "UK-wide"],
                ["NHS prescribing after", "Yes", "Yes (via shared care)", "Only if shared care accepted"],
            ],
            [CONTENT_W * 0.22, CONTENT_W * 0.24, CONTENT_W * 0.27, CONTENT_W * 0.27]),
        Paragraph("A simple decision framework", S["h2"]),
        *bullets([
            "Can’t or don’t want to pay, can wait → NHS standard.",
            "In England, want it free and faster → Right to Choose.",
            "Can pay, want speed, GP likely to accept shared care → private (check first).",
            "Unsure → ask your GP for the local NHS wait, then decide with real numbers.",
        ]),
        PageBreak(),
    ]
    return f


PROVIDER_CRITERIA = [
    "Assessment cost", "Titration cost (per appt & typical total)",
    "Prescription cost during titration", "Annual review cost",
    "Current wait: assessment", "Current wait: treatment after diagnosis",
    "Right to Choose available?", "Format (video / in person)",
    "Assessment length & structure", "Who assesses (psychiatrist / nurse)",
    "GMC / NMC registration checked?", "CQC registered?",
    "Diagnostic interview (DSM-5 / ICD-11)?", "Informant questionnaire used?",
    "Titration offered?", "Shared care supported?",
    "Shared care acceptance rate with GPs", "What if GP refuses shared care?",
    "Support between appointments", "Cancellation / rebooking policy",
    "Patient reviews (after diagnosis)", "Gut feeling after contact",
]


def sec_provider_comparison(num, total):
    f = section_header(num, total, "Provider Comparison Template",
        "22 criteria, three providers, side by side. Fill it in as you research — "
        "the decision usually makes itself.")
    f += note_box("Right to Choose warning: if you’re in England, check whether each provider "
                  "offers RTC before you pay anything — the same assessment may be free.")
    rows = [[c, "", "", ""] for c in PROVIDER_CRITERIA]
    f += grid_table(
        ["Criteria", "Provider 1:", "Provider 2:", "Provider 3:"],
        rows,
        [CONTENT_W * 0.34, CONTENT_W * 0.22, CONTENT_W * 0.22, CONTENT_W * 0.22],
        row_h=18)
    f += [PageBreak()]
    return f


def sec_referral_tracker(num, total):
    f = section_header(num, total, "Referral Status Tracker",
        "A referral only moves if someone is watching it. This page is the someone.")
    f += [
        Paragraph("Referral details", S["h2"]),
        *grid_table(
            ["Field", "Details"],
            [["Referred by (GP / surgery)", ""], ["Date sent", ""],
             ["Referred to (provider)", ""], ["Reference number", ""],
             ["Estimated wait (as quoted)", ""], ["Route (NHS / RTC / private)", ""]],
            [CONTENT_W * 0.35, CONTENT_W * 0.65], row_h=20),
        Paragraph("Milestones", S["h2"]),
        *grid_table(
            ["Milestone", "Expected", "Actual", "Notes"],
            [["Referral confirmed received", "", "", ""],
             ["Screening questionnaires sent", "", "", ""],
             ["Questionnaires returned", "", "", ""],
             ["Assessment date offered", "", "", ""],
             ["Assessment attended", "", "", ""],
             ["Report received", "", "", ""]],
            [CONTENT_W * 0.34, CONTENT_W * 0.18, CONTENT_W * 0.18, CONTENT_W * 0.30],
            row_h=20),
        Paragraph("Contact log (every call, letter and email — dated)", S["h2"]),
        *grid_table(
            ["Date", "Who", "What was said", "Next step"],
            [["", "", "", ""] for _ in range(9)],
            [CONTENT_W * 0.13, CONTENT_W * 0.2, CONTENT_W * 0.37, CONTENT_W * 0.30],
            row_h=22),
        *note_box("Chase every 6–8 weeks. Set the next reminder the moment you finish each call."),
        PageBreak(),
    ]
    return f


ADMIN_TASKS = [
    ("Before the GP", [
        "Book GP appointment (double if possible)",
        "Complete GP prep template (Parts A–H)",
        "Fill in symptom evidence builder",
        "Complete childhood & adult worksheet",
        "Find old school reports (if they exist)",
        "Practise opening statement once, out loud",
    ]),
    ("Referral stage", [
        "Confirm referral sent (date + destination)",
        "Get a copy of the referral letter",
        "Check Right to Choose position for your area (England)",
        "Set first chase-up reminder (6–8 weeks)",
        "Start the referral status tracker",
        "Return screening questionnaires promptly",
    ]),
    ("While waiting", [
        "Keep adding to the evidence sheets",
        "Chase referral on schedule; log every contact",
        "Set up the weekly survival planner",
        "Warn your childhood informant they may get a questionnaire",
    ]),
    ("Assessment & after", [
        "Complete assessment day prep worksheet",
        "Attend assessment; run follow-up checklist after",
        "Request and file the assessment report",
        "Complete “what to ask after assessment” questions",
    ]),
    ("Treatment stage", [
        "Start medication & titration notes (if prescribed)",
        "Ask about shared care before titration ends",
        "Book and prep for each review appointment",
    ]),
]


def sec_admin_tracker(num, total):
    f = section_header(num, total, "Admin Tracker",
        "Your whole process as one checklist, in order. Tick things off; "
        "add your own at the bottom of each stage.")
    for stage, tasks in ADMIN_TASKS:
        f += [Paragraph(stage, S["h2"]), *checks(tasks)]
    f += [*write_lines(3, "Your additions:"), PageBreak()]
    return f


def sec_weekly_planner(num, total):
    f = section_header(num, total, "Weekly Survival Planner",
        "For the wait between referral and assessment. The goal is not a productive week — "
        "it’s a survivable one that doesn’t generate new crises.")
    f += [
        Paragraph("The minimum viable week", S["h2"]),
        *bullets([
            "One sleep anchor: a consistent-ish wake time beats a perfect bedtime.",
            "One priority per day — not three. Write it the night before.",
            "One admin slot (30 minutes, timer on) for the scariest envelope or email.",
            "One thing you’re allowed to drop this week, on purpose, guilt-free.",
        ]),
        Paragraph("This week", S["h2"]),
        *grid_table(
            ["Day", "The one priority", "Done?"],
            [[d, "", ""] for d in ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]],
            [CONTENT_W * 0.15, CONTENT_W * 0.7, CONTENT_W * 0.15], row_h=22),
        Paragraph("Monthly check-in (once, honestly)", S["h2"]),
        *checks([
            "Has the referral been chased on schedule?",
            "Is anything new worth adding to the evidence sheets?",
            "Is my mental health holding? (If not — that’s a GP matter now, not after the assessment.)",
        ]),
        *write_lines(3, "Notes to self:"),
        PageBreak(),
    ]
    return f


def sec_assessment_day(num, total):
    f = section_header(num, total, "Assessment Day Prep Worksheet",
        "The assessment is long (usually 1–3 hours) and thorough. You can’t fail it. "
        "Your only job is to be honest and bring your evidence.")
    f += [
        Paragraph("What to expect", S["h2"]),
        *bullets([
            "A structured diagnostic interview about your life now and in childhood.",
            "Questionnaires (sometimes sent in advance — return them promptly).",
            "Possibly an informant interview or questionnaire about your childhood.",
            "Honest answers beat impressive ones. Bad days are data too — don’t mask.",
        ]),
        Paragraph("Logistics", S["h2"]),
        *grid_table(
            ["Field", "Details"],
            [["Date & time", ""], ["Format / location / link", ""],
             ["Tech check done (if video)", ""], ["Documents to bring / upload", ""],
             ["Time booked off work", ""]],
            [CONTENT_W * 0.35, CONTENT_W * 0.65], row_h=20),
        Paragraph("The five things I must not forget to say", S["h2"]),
        *write_lines(5),
        Paragraph("Questions to ask at the end", S["h2"]),
        *checks([
            "When and how will I get the report?",
            "If diagnosed — what happens next, and what’s the wait for treatment?",
            "If not diagnosed — what explanations should be explored instead?",
            "Who do I contact if I have questions once the report arrives?",
        ]),
        *write_lines(3, "Notes straight after (same day):"),
        PageBreak(),
    ]
    return f


def sec_after_assessment(num, total):
    f = section_header(num, total, "What to Ask After Assessment",
        "Two scenarios, both with sensible next moves. Either way: get the report, "
        "read it twice, and keep it filed.")
    f += [
        Paragraph("If you were diagnosed", S["h2"]),
        *checks([
            "What subtype/presentation was recorded, and what severity?",
            "What are the treatment options — medication, coaching, therapy, adjustments?",
            "If medication: what’s the titration plan, and the wait to start?",
            "Will you send a shared-care agreement to my GP once I’m stable?",
            "What should I tell my employer/university, and is a letter available?",
        ]),
        Paragraph("If you were not diagnosed", S["h2"]),
        *checks([
            "What explanations fit the difficulties better — and what’s the evidence?",
            "Was childhood evidence the gap? Could an informant change the picture?",
            "Can I see exactly which criteria were and weren’t met?",
            "What support exists for the difficulties regardless of label?",
            "Is a second opinion reasonable, and how would that work?",
        ]),
        Paragraph("Reading the report", S["h2"]),
        *bullets([
            "Check the facts: history, dates, quotes. Factual errors can be corrected — ask.",
            "Find the recommendations section — that’s the part your GP acts on.",
            "If anything is unclear, you’re entitled to ask the assessor to explain it.",
        ]),
        *write_lines(3, "My questions for the follow-up:"),
        PageBreak(),
    ]
    return f


def sec_shared_care(num, total):
    f = section_header(num, total, "Shared Care Question Sheet",
        "Shared care moves routine prescribing to your NHS GP after titration. "
        "Ask these early — the answers change which provider makes sense.")
    f += [
        Paragraph("Ask your GP surgery", S["h2"]),
        *checks([
            "Do you accept shared-care agreements for adult ADHD medication?",
            "Are there providers you already have agreements with?",
            "Any you won’t accept (e.g. private-only clinics)?",
            "If you decline, can I have the reason in writing?",
        ]),
        Paragraph("Ask the provider", S["h2"]),
        *checks([
            "How often do GP surgeries actually accept your shared-care agreements?",
            "At what point in titration do you send the agreement?",
            "If my GP refuses — what are my options and ongoing costs with you?",
            "Who handles prescriptions in the gap between titration and shared care starting?",
        ]),
        Paragraph("If shared care is refused", S["h2"]),
        *bullets([
            "Get the surgery’s reason in writing.",
            "Ask the provider to contact the surgery directly — some will.",
            "Budget for private prescribing while you decide (switching GP practice is "
            "sometimes an option — ask about the new surgery’s policy before moving).",
        ]),
        *write_lines(3, "Answers / notes:"),
        PageBreak(),
    ]
    return f


def sec_medication_log(num, total):
    f = section_header(num, total, "Medication & Titration Notes",
        "One minute a day. Rate 1–5 (1 = rough, 5 = great). Your reviewer can only "
        "work with the data you bring.")
    f += [
        Paragraph("Current details", S["h2"]),
        *grid_table(
            ["Field", "Details"],
            [["Medication & dose", ""], ["Time(s) taken", ""],
             ["Started on", ""], ["Prescriber & contact", ""],
             ["Next review date", ""]],
            [CONTENT_W * 0.35, CONTENT_W * 0.65], row_h=20),
    ]
    for week in range(1, 5):
        f += [
            Paragraph(f"Week {week}", S["h2"]),
            *grid_table(
                ["Day", "Focus", "Mood", "Energy", "Appetite", "Sleep", "Side effects / notes"],
                [[d, "", "", "", "", "", ""] for d in
                 ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]],
                [CONTENT_W * 0.10, CONTENT_W * 0.09, CONTENT_W * 0.09, CONTENT_W * 0.09,
                 CONTENT_W * 0.10, CONTENT_W * 0.09, CONTENT_W * 0.44],
                row_h=17),
        ]
        if week == 2:
            f.append(PageBreak())
    f += [
        Paragraph("Questions for my next review", S["h2"]),
        *write_lines(3),
        *note_box("Never change a dose except as directed by your prescriber. Concerning side "
                  "effects — chest pain, severe mood changes, thoughts of self-harm — mean "
                  "contacting your prescriber or NHS 111 promptly; 999 in an emergency."),
        PageBreak(),
    ]
    return f


def closing(upsell_para):
    return [
        Paragraph("The small print", S["h2"]),
        *disclaimer_flow(),
        Spacer(1, 10),
        HRFlowable(width="100%", thickness=1.2, color=CORAL, spaceAfter=10),
        Paragraph(upsell_para, S["body"]),
        Paragraph("Made with patience, for people who lost theirs.", S["small"]),
    ]


# ===========================================================================
# PDF assemblies
# ===========================================================================

def build_starter_pdf(path):
    T = 7
    story = cover(
        "Diagnosis Prep<br/>Starter Pack",
        "From “I think I have ADHD” to a GP referral",
        "Your step-by-step system for the first stretch of the journey — "
        "without losing your mind in the process.",
        ["Full GP appointment prep template (Parts A–H)",
         "Full symptom evidence builder (8 categories)",
         "Childhood & adult examples worksheet",
         "Phone scripts: every scenario (9 scripts)",
         "Appointment follow-up checklist",
         "AI prompt pack (9 prompts)",
         "Referral document checklist"],
        f"UK edition · Starter Pack · {SITE}")
    story += sec_gp_prep(1, T)
    story += sec_evidence(2, T)
    story += sec_childhood(3, T)
    story += sec_scripts(4, T, include_rtc=False)
    story += sec_followup(5, T)
    story += sec_ai_prompts(6, T, full=False)
    story += sec_referral_docs(7, T)
    story += closing(
        "Going further? The Full Diagnosis Prep & Survival Pack (£24) adds the route guide, "
        "provider comparison, referral & admin trackers, weekly survival planner, assessment "
        f"day prep, shared care questions, titration log and a 20-prompt AI library — at {SITE}/full-pack")
    build_pdf(path, "Diagnosis Prep Starter Pack — ADHD Assist",
              f"Diagnosis Prep Starter Pack  ·  {SITE}", story)


def build_full_pdf(path):
    T = 15
    story = cover(
        "Full Diagnosis Prep<br/>&amp; Survival Pack",
        "From suspicion to titration — the complete system",
        "Everything from first GP appointment to medication reviews, so you never "
        "have to hold the whole process in your head.",
        ["Understanding your routes (with comparison table)",
         "Full GP prep template · symptom evidence builder",
         "Childhood & adult worksheet · 10 phone scripts",
         "Provider comparison · referral & admin trackers",
         "Weekly survival planner · assessment day prep",
         "After-assessment guide · shared care questions",
         "Medication & titration notes · 20-prompt AI library"],
        f"UK edition · Full Pack · {SITE}")
    story += sec_routes_full(1, T)
    story += sec_gp_prep(2, T)
    story += sec_evidence(3, T)
    story += sec_childhood(4, T)
    story += sec_scripts(5, T, include_rtc=True)
    story += sec_provider_comparison(6, T)
    story += sec_referral_tracker(7, T)
    story += sec_admin_tracker(8, T)
    story += sec_weekly_planner(9, T)
    story += sec_assessment_day(10, T)
    story += sec_followup(11, T)
    story += sec_after_assessment(12, T)
    story += sec_shared_care(13, T)
    story += sec_medication_log(14, T)
    story += sec_ai_prompts(15, T, full=True)
    story += closing(
        "Want a live tracking system too? The ADHD Admin System (£12, or £35 bundled with this "
        f"pack) is an interactive tracker — browser app, spreadsheet and Notion guide — at {SITE}/admin-system")
    build_pdf(path, "Full Diagnosis Prep & Survival Pack — ADHD Assist",
              f"Full Diagnosis Prep & Survival Pack  ·  {SITE}", story)


# ===========================================================================
# Admin System: Excel workbook + Notion guide (HTML app is a template file)
# ===========================================================================

def build_admin_xlsx(path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    INK_HEX, CORAL_HEX, CREAM_HEX = "FF1F5F6B", "FFFF6B6B", "FFF5F1E8"
    head_fill = PatternFill("solid", fgColor=INK_HEX)
    head_font = Font(bold=True, color="FFFFFFFF", size=10)
    title_font = Font(bold=True, color=INK_HEX, size=14)
    thin = Side(style="thin", color="FF9DBEC5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    def sheet(name, title, headers, widths, rows, status_col=None, status_opts=None):
        ws = wb.create_sheet(name)
        ws["A1"] = title
        ws["A1"].font = title_font
        ws.append([])
        ws.append(headers)
        hr = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=hr, column=c)
            cell.fill = head_fill
            cell.font = head_font
            cell.border = border
            ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
        for r in rows:
            ws.append(r)
        for row in ws.iter_rows(min_row=hr + 1, max_row=ws.max_row,
                                max_col=len(headers)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        if status_col and status_opts:
            dv = DataValidation(type="list",
                                formula1='"' + ",".join(status_opts) + '"',
                                allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{status_col}{hr + 1}:{status_col}{max(ws.max_row, hr + 40)}")
        return ws

    # Dashboard
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = "ADHD Admin System — Dashboard"
    ws["A1"].font = title_font
    ws["A3"] = "Milestones (mark Done as you go — the tracker tabs hold the detail)"
    ws["A3"].font = Font(bold=True, color=INK_HEX)
    milestones = ["GP appointment booked", "GP appointment attended", "Referral sent",
                  "Referral confirmed", "Questionnaires returned", "Assessment date offered",
                  "Assessment attended", "Report received", "Treatment started",
                  "Shared care agreed"]
    ws.append([]); ws.append(["Milestone", "Status", "Date", "Notes"])
    hr = ws.max_row
    for c, w in zip(range(1, 5), [34, 14, 14, 40]):
        cell = ws.cell(row=hr, column=c)
        cell.fill, cell.font, cell.border = head_fill, head_font, border
        ws.column_dimensions[get_column_letter(c)].width = w
    for m in milestones:
        ws.append([m, "Not started", "", ""])
    for row in ws.iter_rows(min_row=hr + 1, max_row=ws.max_row, max_col=4):
        for cell in row:
            cell.border = border
    dv = DataValidation(type="list", formula1='"Not started,In progress,Done"',
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"B{hr + 1}:B{ws.max_row}")
    ws[f"A{ws.max_row + 2}"] = "Stuck? Do the first not-started milestone. That’s always the answer."
    ws[f"A{ws.max_row}"].font = Font(italic=True, color=INK_HEX)

    # Admin Tracker
    tasks = [(stage, task) for stage, ts in ADMIN_TASKS for task in ts]
    sheet("Admin Tracker", "Admin Tracker — the whole process, one list",
          ["Stage", "Task", "Status", "Due / done", "Notes"],
          [18, 46, 14, 14, 36],
          [[s, t, "Not started", "", ""] for s, t in tasks],
          status_col="C", status_opts=["Not started", "In progress", "Done"])

    # Referral Log
    sheet("Referral Log", "Referral Log — details + every contact, dated",
          ["Date", "Who / organisation", "What was said", "Next step + who owes it"],
          [13, 26, 44, 34],
          [["", "", "", ""] for _ in range(25)])

    # Provider Compare
    sheet("Provider Compare", "Provider Comparison — 22 criteria, 3 providers",
          ["Criteria", "Provider 1", "Provider 2", "Provider 3"],
          [42, 24, 24, 24],
          [[c, "", "", ""] for c in PROVIDER_CRITERIA])

    # Documents
    docs = [("Before GP", d) for d in
            ["Symptom evidence sheets", "School reports / childhood records",
             "Current medication list", "Medical history notes"]] + \
           [("Referral stage", d) for d in
            ["Copy of referral letter", "Referral reference number",
             "Screening questionnaires (returned)", "Photo ID / NHS number"]] + \
           [("Assessment & after", d) for d in
            ["Assessment report", "Diagnosis letter (if applicable)",
             "Shared care agreement", "Prescriptions / titration schedule"]]
    sheet("Documents", "Documents — pre-listed across 3 stages",
          ["Stage", "Document", "Priority", "Status", "Where it lives"],
          [18, 36, 12, 14, 30],
          [[s, d, "High", "Not started", ""] for s, d in docs],
          status_col="D", status_opts=["Not started", "In progress", "Done"])

    # Appointment Log
    sheet("Appointment Log", "Appointment Log — add a row per appointment",
          ["Date", "Type", "Who with", "What was agreed", "Next step", "How it felt"],
          [13, 16, 20, 34, 26, 22],
          [["", "", "", "", "", ""] for _ in range(20)],
          status_col="B",
          status_opts=["GP", "Assessment", "Titration", "Review", "Other"])

    # Medication
    med_rows = []
    for week in range(1, 5):
        med_rows.append([f"Week {week}", "", "", "", "", "", ""])
        for day in ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]:
            med_rows.append([day, "", "", "", "", "", ""])
    sheet("Medication", "Medication — 14+ day daily log (rate 1–5)",
          ["Day", "Focus", "Mood", "Energy", "Appetite", "Sleep", "Side effects / notes"],
          [12, 9, 9, 9, 10, 9, 44], med_rows)

    # How to Use
    ws = wb.create_sheet("How to Use")
    ws["A1"] = "How to Use"
    ws["A1"].font = title_font
    ws.column_dimensions["A"].width = 100
    tips = [
        "", "Each tab covers one stage of the process. Start at the Dashboard.",
        "Status cells have dropdowns — click a cell and pick from the list.",
        "Add rows anywhere; the formatting copies down if you insert within the table.",
        "Log EVERY contact in the Referral Log — dates win arguments with waiting lists.",
        "The web-app version of this system (in the same download) saves automatically in your browser.",
        "", "Educational only — see the disclaimer in your pack and on the website. Not medical advice.",
    ]
    for t in tips:
        ws.append([t])
    wb.save(path)


def build_notion_guide(path):
    story = [
        Paragraph("ADHD Admin System", S["cover_title"]),
        Paragraph("Notion Workspace Build Guide", S["cover_sub"]),
        Paragraph(
            "Prefer Notion? This guide rebuilds the whole Admin System as a Notion workspace "
            "in about 20 minutes. You’ll end up with the same eight views as the app and "
            "spreadsheet, with Notion’s reminders and mobile app on top.", S["body"]),
        PageBreak(),
        *section_header(1, 4, "Set up the shell"),
        *bullets([
            "Create a new page called “ADHD Admin System”. Give it an icon you like seeing.",
            "Add a Callout at the top: “One list. One log. Chase every 6–8 weeks.”",
            "Everything below lives inside this one page — resist making it fancy.",
        ]),
        *section_header(2, 4, "Create the two databases"),
        Paragraph("Database 1 — “Tasks & Milestones” (table)", S["h2"]),
        *bullets([
            "Properties: Name (title) · Stage (select: Before GP / Referral / Waiting / "
            "Assessment / Treatment) · Status (select: Not started / In progress / Done) · "
            "Due (date) · Notes (text).",
            "Copy the tasks in from the Admin Tracker tab of the spreadsheet (paste as new rows).",
            "Add a Board view grouped by Status — this becomes your dashboard.",
            "Add a filtered view “Next up”: Status is Not started, sorted by Stage.",
        ]),
        Paragraph("Database 2 — “Contact Log” (table)", S["h2"]),
        *bullets([
            "Properties: What was said (title) · Date (date) · Who (text) · "
            "Next step (text) · Owed by (select: Me / Them).",
            "Sort by Date, newest first. Log every call, letter, and email here.",
            "Add a reminder on the Date property when the next chase-up is due.",
        ]),
        *section_header(3, 4, "Recreate the remaining tabs"),
        *bullets([
            "Provider Compare: a simple table block with the 22 criteria from the spreadsheet "
            "as rows and three provider columns.",
            "Documents: a checklist (to-do blocks) grouped under three headings — "
            "Before GP / Referral stage / Assessment & after.",
            "Appointment Log: either a third database (copy Contact Log, add a Type select: "
            "GP / Assessment / Titration / Review) or extra rows in Contact Log.",
            "Medication: a table block with Day / Focus / Mood / Energy / Appetite / Sleep / "
            "Notes columns — duplicate it for each titration week.",
        ]),
        *section_header(4, 4, "Make it ADHD-proof"),
        *bullets([
            "Pin the page to your sidebar favourites and phone home screen.",
            "Set ONE recurring reminder: “Open the Admin System” — weekly, same day, same time.",
            "When you don’t know what to do next: open “Next up”, do the top item only.",
            "Don’t redesign the system when overwhelmed. Using an ugly system beats "
            "building a beautiful one.",
        ]),
        Spacer(1, 10),
        *note_box("Educational only — this system organises your process; it is not medical "
                  "advice. See the full disclaimer in your pack."),
    ]
    build_pdf(path, "ADHD Admin System — Notion Build Guide — ADHD Assist",
              f"ADHD Admin System · Notion Build Guide  ·  {SITE}", story)


ADMIN_README = """ADHD Admin System — what's in this download
============================================

1. ADHD-Admin-System.html
   The interactive tracker. Open it in any browser (double-click it).
   Everything you type saves automatically in that browser via localStorage.
   No account, no internet needed after opening.
   NOTE: data lives in the browser profile on that device. Use the Export
   button (top right) to back up, and Import to restore or move devices.

2. ADHD-Admin-System.xlsx
   The same system as a spreadsheet, for Excel / Google Sheets / LibreOffice.
   To use in Google Sheets: drive.google.com → New → File upload → open with
   Google Sheets.

3. Notion-Build-Guide.pdf
   Step-by-step instructions to rebuild the system as a Notion workspace.

Educational only — not medical advice. Full disclaimer on the website.
"""


# ===========================================================================
# Main
# ===========================================================================

def zip_files(zip_path, files):
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for src, arcname in files:
            z.write(src, arcname)
    print(f"Built {zip_path}")


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    tmp = os.path.join(OUT_DIR, "_tmp")
    os.makedirs(tmp, exist_ok=True)

    starter_pdf = os.path.join(tmp, "Diagnosis-Prep-Starter-Pack.pdf")
    full_pdf = os.path.join(tmp, "Full-Diagnosis-Prep-and-Survival-Pack.pdf")
    xlsx = os.path.join(tmp, "ADHD-Admin-System.xlsx")
    notion = os.path.join(tmp, "Notion-Build-Guide.pdf")
    html_src = os.path.join(HERE, "templates", "ADHD-Admin-System.html")
    readme = os.path.join(tmp, "README.txt")

    build_starter_pdf(starter_pdf)
    build_full_pdf(full_pdf)
    build_admin_xlsx(xlsx)
    build_notion_guide(notion)
    with open(readme, "w") as fh:
        fh.write(ADMIN_README)

    zip_files(os.path.join(OUT_DIR, "starter-pack.zip"),
              [(starter_pdf, "Diagnosis-Prep-Starter-Pack.pdf")])
    zip_files(os.path.join(OUT_DIR, "full-pack.zip"),
              [(full_pdf, "Full-Diagnosis-Prep-and-Survival-Pack.pdf")])
    admin_files = [
        (html_src, "ADHD-Admin-System.html"),
        (xlsx, "ADHD-Admin-System.xlsx"),
        (notion, "Notion-Build-Guide.pdf"),
        (readme, "README.txt"),
    ]
    zip_files(os.path.join(OUT_DIR, "admin-system.zip"), admin_files)
    zip_files(os.path.join(OUT_DIR, "bundle.zip"),
              [(full_pdf, "Full-Diagnosis-Prep-and-Survival-Pack.pdf")] + admin_files)

    shutil.rmtree(tmp)


if __name__ == "__main__":
    main()
